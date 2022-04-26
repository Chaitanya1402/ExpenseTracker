# AUTOMATE EXPENSES BY DATE AND AMOUNT

# Use command prompt
# python exp.py particulars go here +- amount

import openpyxl
from openpyxl.styles import Font, Alignment
import datetime
import os
import sys


def new_sheet_create():                               # Called at the beginning of each month
    print("Manually merge cells of last day of the previous month")
    global sheet
    sheet = book.create_sheet(now.strftime("%B %Y"))  # Create a new sheet of required form
    if "Sheet" in book.sheetnames:                    # Delete any unwanted sheet if exists
        del book["Sheet"]
    sheet.column_dimensions["A"].width = 15           # Set the column width and contents
    sheet["A1"].value = "DATE"
    sheet.column_dimensions["B"].width = 40           # Do for all columns
    sheet["B1"].value = "PARTICULARS"
    sheet.column_dimensions["C"].width = 13
    sheet["C1"].value = "DEBIT"
    sheet.column_dimensions["D"].width = 13
    sheet["D1"].value = "CREDIT"
    sheet.column_dimensions["E"].width = 13
    sheet["E1"].value = "SAVINGS"
    for col in range(1, 6):                           # To add some modifications
        var = sheet.cell(row=1, column=col)           # Select 1st row
        var.font = Font(name="Franklin Gothic Demi", size=12)  # Set font size and style
        var.alignment = Alignment(horizontal="center", vertical="center")  # Set alignment


def add(tot):                # Active every time program is executed
    # If last row value does not match the date, then call merge_cells() function
    if now.strftime("%d-%m-%Y") != sheet.cell(row=sheet.max_row, column=1).value and sheet.max_row != 1:
        to_merge_cells()
    row = sheet.max_row+1    # Go to next row
    var = sheet.cell(row=row, column=1)   # First column
    var.value = now.strftime("%d-%m-%Y")  # First column date
    var.alignment = Alignment(vertical="center", horizontal="left")  # Align
    sheet.cell(row=row, column=2).value = " ".join(sys.argv[1:-1]).title()  # Second column from cmd
    try:                                  # If amount is not provided, handle exception
        if int(sys.argv[-1]) < 0:            # Negative for debit
            sheet.cell(row=row, column=3).value = abs(int(sys.argv[-1]))  # Store in third column
        else:                             # Positive for credit
            sheet.cell(row=row, column=4).value = int(sys.argv[-1])  # Store in fourth column
        tot += int(sys.argv[-1])          # Increment total
    except ValueError:                    # Exception handling
        print("\nUse correct syntax:")
        print("exp <particulars> <+-amount>")
        print("+ for Credit (Income)\n- for Debit (Spent)")
        sys.exit()
    sheet.cell(row=row, column=5).value = tot  # Store in fifth column
    # You provide negative value (ex: -540) for debit, but a positive value will be stored in debit
    # column and negative value is added to last column
    for col in range(1, 6):               # Use for Font modifications
        var = sheet.cell(row=row, column=col)
        var.font = Font(name="Cambria", size=12)  # Set font size and type


def to_merge_cells():   # Called when date format does not match previous row date
    i = 0               # Do this because the last entered date may not always be the previous
    while True:         # date. It may be 2 days ago as well (If you didn't enter any data for last day)
        i += 1          # Temp variable to store days time span
        day_ago = datetime.timedelta(days=i)
        final = now - day_ago  # Set the new datetime object
        if final.strftime("%d-%m-%Y") == sheet.cell(row=sheet.max_row, column=1).value:
            # If date matches then break
            # Get contents of last row 1st column
            break
    """ Previous if statement gets you the last occurrence of last date entered before today.
     Next if statement gets you first occurrence of last date entered before today
     So merge cells from first occurrence to last one
     So even if you enter only one data for a given day, then first and last occurrence would
     be the same cell. So even if you merge the cell, nothing would happen
     Which is what you want.
    """
    for row_num in range(2, sheet.max_row+1):
        if sheet.cell(row=row_num, column=1).value == final.strftime("%d-%m-%Y"):
            # Get first occurrence
            sheet.merge_cells("A{}:A{}".format(row_num, sheet.max_row))
            # print("cells merged")
            break


now = datetime.datetime.now()               # Get the current date
loc_folder = "C:\\Users\\admin\\PycharmProjects\\HelloWorld"         # CHANGE THIS
file_name = "expenses" + str(now.year) + ".xlsx"  # File name is expenses(YEAR).xlsx
if file_name not in os.listdir(loc_folder):  # If file not exists,
    book = openpyxl.Workbook()               # Create a new one
else:                                        # If it exists, open it
    book = openpyxl.load_workbook(loc_folder + "\\expenses{}.xlsx".format(now.year))
if now.strftime("%B %Y") not in book.sheetnames:  # Check for sheet names
    total = 0                                # Total only for last column
    new_sheet_create()                       # If not present, create a new one (call function).
else:
    sheet = book[now.strftime("%B %Y")]      # Else open the one which matches criteria
    total = sheet.cell(row=sheet.max_row, column=5).value     # Get last entered value in total
# print(loc_folder + "\\expenses{}.xlsx".format(now.year))
add(total)                                   # Call function to add the expenses
try:
    book.save(loc_folder + "\\expenses{}.xlsx".format(now.year))  # Once returned, save it in the same file
except PermissionError:
    print("ERROR: Close the workbook first")
book.close()                                 # Close the book
